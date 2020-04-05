#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Fri Apr  3 04:12:42 2020

@author: tiago
"""

import math
import locale
from openpyxl import load_workbook
path = "../Data/"

def getData2(ini, end):
    doc    = "testPoints.xlsx"
    wbk     = load_workbook(path+doc)
    sheet   = wbk["Experiment2"]
    cells   = sheet[ini : end]
    
    position = []
    rss = []

    i = -1
    locale.setlocale(locale.LC_ALL, 'en_US.UTF-8')
    for pos1, pos2, c1, c2, c3, c4, c5, c6 in cells:
        i = i + 1
        position.append([])
        position[i].append(pos1.value)
        position[i].append(pos2.value)
        rss.append([c1.value, c2.value, c3.value, c4.value, c5.value, c6.value])

    return rss, position

def storeData(data,sheetName):
    doc    = "testPoints.xlsx"
    wbk     = load_workbook(path+doc)
    page    = wbk[sheetName]
    page.append(data)
    wbk.save(path+doc)

def dist_rss(pos, anchors, rss):
    
    data = []
    for anchor in anchors:
        data.append(math.sqrt((pos[0] - anchor[0])**2 + (pos[1] - anchor[1])**2))
    data.extend(rss)
    storeData(data, "dist_rss")

def average_exp2(position, rss):
    
    data = []
    data.extend(position)
    data.extend(rss)
    storeData(data, "Average_exp2")

def main():
    
    anchors = [[-4.0, 4.6],[4.0, 2.4],[-4.6, 1.84],[-2.0, 5.52],[-2.74, 0.92],[2.0, 0.8]]
    rss, position = getData2("A2", "H18")
    
    for i in range(17):
        average_exp2(position[i], rss[i])
    
if __name__== "__main__":
    main()