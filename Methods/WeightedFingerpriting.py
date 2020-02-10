#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Sun Feb  9 18:54:06 2020

@author: tiago
"""


import sys
import time
import math
import locale
from openpyxl import load_workbook

x           = [4.4, 2.2, 0, 6.6, 0, 6.6]
y           = [2.6,13.0,6.5, 10.4, 3.9, 7.8]
h           = [0.76, 1.7, 1.65, 1.17, 2.02, 1.52]


def storeData(data,sheetName):
    path    = "methods.xlsx"
    wbk     = load_workbook(path)
    page    = wbk[sheetName]
    page.append(data)
    wbk.save(path)

def getData(path, page, ini, end):
    wbk     = load_workbook(path)
    sheet   = wbk[page]
    cells   = sheet[ini : end]
    
    position = []
    row = []

    i = -1
    locale.setlocale(locale.LC_ALL, 'en_US.UTF-8')
    for pos1, pos2, c1, c2, c3, c4, c5, c6 in cells:
        i = i + 1
        position.append([])
        position[i].append(locale.atof(pos1.value))
        position[i].append(locale.atof(pos2.value))
        row.append([locale.atof(c1.value), locale.atof(c2.value), locale.atof(c3.value), locale.atof(c4.value), locale.atof(c5.value)])
    
    return row, position

def main():
    
    n = int(sys.argv[1])
    
    # get list of rssi and list of position, respectivelly
    l_base, base_pos = getData('dataBase.xlsx',"Average" ,"A2", "H27")
    l_test, test_pos = getData('testPoints.xlsx', "Average",  "A2", "H19")
    
    k = -1
    w0 = 10**(-8)
    for point in l_test:
        k = k + 1
        cost = []
        # execute Fingerpriting and calculate its processing time
        start = time.time()
        for i in range(26):
            cost.append((point[0] - l_base[i][0])**2 + (point[1] - l_base[i][1])**2 + (point[2] - l_base[i][2])**2 + (point[3] - l_base[i][3])**2 + (point[4] - l_base[i][4])**2)
        # sort cost and index of base_pos accordingly
        index = sorted(range(len(cost)), key=lambda k: cost[k])
        # set the weight
        weight = 0
        for j in range(0,n):
            weight = weight + 1.0/(cost[index[j]]+w0)
        # calculate x,y via weighted cost
        x = y = 0
        for j in range(0,n):
            x = x + ((1.0/(cost[index[j]]+w0))/weight)*base_pos[index[j]][0]
            y = y + ((1.0/(cost[index[j]]+w0))/weight)*base_pos[index[j]][1]
        duration = time.time() - start
        estimate = [round(x, 2),round(y, 2)]
        # calculate error
        error = math.sqrt((estimate[0] - test_pos[k][0])**2 + (estimate[1] - test_pos[k][1])**2)
        # put them together
        data = [test_pos[k][0], test_pos[k][1]]
        data.extend(estimate)
        data.append(duration)
        data.append(error)
        # store in xlsx
        storeData(data, "WeightedFingerKnn"+str(n))
        
        
if __name__== "__main__":
        main()