#!/usr/bin/env python2
# -*- coding: utf-8 -*-
"""
Created on Tue Feb  4 13:45:22 2020

@author: mister-c
"""

import time
import math
import locale
import numpy as np
from numpy.linalg import inv
from openpyxl import load_workbook

x           = [4.4, 2.2, 0, 6.6, 0, 6.6]
y           = [2.6,13.0,6.5, 10.4, 3.9, 7.8]
h           = [0.76, 1.7, 1.65, 1.17, 2.02, 1.52]
d           = []
l_rssi      = []
position    = []

def storeData(data,sheetName):
    path    = "methods.xlsx"
    wbk     = load_workbook(path)
    page    = wbk[sheetName]
    page.append(data)
    wbk.save(path)

def findDistance(a, n, rssi):
    return 10**((a - rssi)/(10*n))

# get position and RSSI from excel
def getTestData(path):
    wbk     = load_workbook(path)
    sheet   = wbk["Sheet1"]
    cells   = sheet['A2': 'H181']
    
    for k in range(6):
        l_rssi.append([])
    
    i = -1
    locale.setlocale(locale.LC_ALL, 'en_US.UTF-8')
    for pos1, pos2, c1, c2, c3, c4, c5, c6 in cells:
        i = i + 1
        position.append([])
        position[i].append(locale.atof(pos1.value))
        position[i].append(locale.atof(pos2.value))
        l_rssi[0].append(locale.atof(c1.value))
        l_rssi[1].append(locale.atof(c2.value))
        l_rssi[2].append(locale.atof(c3.value))
        l_rssi[3].append(locale.atof(c4.value))
        l_rssi[4].append(locale.atof(c5.value))
        l_rssi[5].append(locale.atof(c6.value))

def hyperbolicAlgorithm(rssi):
    
    d.append(findDistance(-43.9475, 1.4152, rssi[0]) )
    d.append(findDistance(-34.2028, 1.8470, rssi[1]))
    d.append(findDistance(-32.1226, 2.067, rssi[2]))
    d.append(findDistance(-40.3671, 2.0823, rssi[3]))
    d.append(findDistance(-32.8183, 2.0228,rssi[4]))
    #d.append(findDistance(-41.6982, 1.1648, rssi[5]))
    
    A = np.matrix([[ 2*(x[2] - x[0]), 2*(y[2] - y[0]) ], 
                   [ 2*(x[2] - x[1]), 2*(y[2] - y[1]) ],
                   [ 2*(x[2] - x[4]), 2*(y[2] - y[4]) ],
                   [ 2*(x[2] - x[3]), 2*(y[2] - y[3]) ]
                   #[ 2*(x[5] - x[4]), 2*(y[5] - y[4]) ]
                  ], dtype=np.float64)
    b= np.matrix([
                [d[0]**2 - d[2]**2 - x[0]**2 - y[0]**2 + x[2]**2 + y[2]**2],
                [d[1]**2 - d[2]**2 - x[1]**2 - y[1]**2 + x[2]**2 + y[2]**2],
                [d[4]**2 - d[2]**2 - x[4]**2 - y[4]**2 + x[2]**2 + y[2]**2],
                [d[3]**2 - d[2]**2 - x[3]**2 - y[3]**2 + x[2]**2 + y[2]**2],
                #[d[4]**2 - d[4]**2 - x[4]**2 - y[4]**2 + x[4]**2 + y[4]**2]
                ], dtype=np.float64)
    
    AT  = A.transpose()
    AAT = AT.dot(A)
    d.clear()
    try:
        var = (inv(AAT)*AT*b).tolist()
        return [var[0][0], var[1][0]]
    except:
        print("Could not solve since matrix has no inverse.")
        return -1
    

def main():
    
    getTestData('testPoints.xlsx')
    n = len(position)
    
    for i in range(n):
        rssi = [l_rssi[0][i],l_rssi[1][i],l_rssi[2][i],l_rssi[3][i],l_rssi[4][i],l_rssi[5][i]]
        #calculate duration
        start = time.time()
        res = hyperbolicAlgorithm(rssi)
        duration = time.time() - start
        # calculate precision
        precision = math.sqrt( (position[i][0] - res[0])**2 + (position[i][1] - res[1])**2 )
        # put them together
        data = [position[i][0], position[i][1]]
        data.extend(res)
        data.append(duration)
        data.append(precision)
        # store in xlsx
        storeData(data,"HyperbolicAlgo")
    
if __name__== "__main__":
        main()