#!/usr/bin/env python2
# -*- coding: utf-8 -*-
"""
Created on Mon Feb 17 18:18:16 2020

@author: mister-c
"""

from sklearn.neural_network import MLPRegressor
from sklearn import svm
import sys
import time
import math
import numpy as np
import locale
from openpyxl import load_workbook

x           = [4.4, 2.2, 0, 6.6, 0, 6.6]
y           = [2.6,13.0,6.5, 10.4, 3.9, 7.8]
h           = [0.76, 1.7, 1.65, 1.17, 2.02, 1.52]


def storeData(data,sheetName):
    path    = "methods.xlsx"
    wbk     = load_workbook(path)
    page    = wbk[sheetName]
    page.append(data)
    wbk.save(path)

def getData(path, page, ini, end, times):
    wbk     = load_workbook(path)
    sheet   = wbk[page]
    cells   = sheet[ini : end]
    
    Y = []
    X = []

    i = -1
    locale.setlocale(locale.LC_ALL, 'en_US.UTF-8')
    for distance, c1, c2, c3, c4, c5, c6, c7, c8, c9, c10 in cells:
        i = i + 1
        for k in range(times):
            Y.append(distance.value)
        X.append(c1.value) 
        X.append(c2.value)
        X.append(c3.value) 
        X.append(c4.value) 
        X.append(c5.value)
        X.append(c6.value)
        X.append(c7.value)
        X.append(c8.value)
        X.append(c9.value)
        X.append(c10.value)
    
    return X, Y

def main():
    
    X, Y = getData('dataBase.xlsx',"TiagoLocalizacao1" ,"A2", "K44", 10)
    training_X = np.array(X).reshape(1,-1)
    training_Y = np.array(Y).reshape(1,-1)
    reg = MLPRegressor(solver='lbfgs', activation='identity', hidden_layer_sizes=(100,))
    reg.fit(training_X, training_Y)
    
    #test_rssi, test_pos = getData('testPoints.xlsx', "TiagoLocalizacao1",  "A2", "B19")
    #training_X = np.array(base_rssi)
    #training_Y = np.array(base_pos)
    #test_X = np.array(test_rssi)
    #reg = MLPRegressor(solver='lbfgs', activation='identity', hidden_layer_sizes=(100,))
    #reg.fit(training_X, training_Y)
    #for i in range(0,18):
     #   pred_y_test = reg.predict(test_X[i].reshape(1,-1))
     #   error = np.sqrt((pred_y_test[0][0] - test_pos[i][0])**2 + (pred_y_test[0][1] - test_pos[i][1])**2)
        
if __name__== "__main__":
        main()