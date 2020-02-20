#!/usr/bin/env python2
# -*- coding: utf-8 -*-
"""
Created on Thu Feb 20 15:46:09 2020

@author: mister-c
"""

from sklearn.neural_network import MLPRegressor
from sklearn import svm
import sys
import socket
import time
import math
import pickle
import numpy as np
import locale
from openpyxl import load_workbook

x           = [4.4, 2.2, 0, 6.6, 0, 6.6]
y           = [2.6,13.0,6.5, 10.4, 3.9, 7.8]
h           = [0.76, 1.7, 1.65, 1.17, 2.02, 1.52]

path = "../Data/"

def getData(doc, page, ini, end):
    wbk     = load_workbook(path+doc)
    sheet   = wbk[page]
    cells   = sheet[ini : end]
    
    position = []
    row = []

    i = -1
    locale.setlocale(locale.LC_ALL, 'en_US.UTF-8')
    for pos1, pos2, c1, c2, c3, c4, c5 in cells:
        i = i + 1
        position.append([])
        position[i].append(locale.atof(pos1.value))
        position[i].append(locale.atof(pos2.value))
        row.append([locale.atof(c1.value), locale.atof(c2.value), locale.atof(c3.value), locale.atof(c4.value), locale.atof(c5.value)])
    
    return row, position

def storeData(data,sheetName):
    doc    = "methods.xlsx"
    wbk     = load_workbook(path+doc)
    page    = wbk[sheetName]
    page.append(data)
    wbk.save(path+doc)


def fingerNN(reg, test_X, test_pos):
    
    start = time.time()
    pred_y_test = reg.predict(test_X.reshape(1,-1))
    duration = time.time() - start
    error = np.sqrt((pred_y_test[0][0] - test_pos[0])**2 + (pred_y_test[0][1] - test_pos[1])**2)
    data = []
    data.extend(test_pos)
    data.extend(pred_y_test[0])
    data.append(duration)
    data.append(error)
    
    return data

def main():
    
    # next create a socket object 
    s = socket.socket()          
    print("Socket successfully created")  
    # reserve a port on your computer in our 
    # case it is 12345 but it can be anything 
    port = 12225                  
    # Next bind to the port 
    # we have not typed any ip in the ip field 
    # instead we have inputted an empty string 
    # this makes the server listen to requests  
    # coming from other computers on the network 
    s.bind(('', port))         
    print("socket binded to %s" %(port))  
    # put the socket into listening mode 
    s.listen(1)      
    print("socket is listening")            
    # Establish connection with client. 
    channel, addr = s.accept()
    print('Got connection from', addr) 
     
    
    base_rssi, base_pos = getData('dataBase.xlsx',"Sheet1" ,"A2", "G431")
    test_rssi, test_pos = getData('testPoints.xlsx', "Average",  "A2", "G19")
    training_X = np.array(base_rssi)
    training_Y = np.array(base_pos)
    test_X = np.array(test_rssi)
    reg = MLPRegressor(solver='lbfgs', activation='identity', hidden_layer_sizes=(100,))
    reg.fit(training_X, training_Y)
    
    for k in range(0,18):
        recd_data = channel.recv( 1024 )
        recd_data = pickle.loads(recd_data)
        send_data = fingerNN(reg, np.array(recd_data), np.array(test_pos[k]))
        print(send_data)
        channel.send(pickle.dumps(send_data))
    channel.close()
        
if __name__== "__main__":
        main()