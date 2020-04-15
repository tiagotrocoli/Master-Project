#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Wed Apr 15 09:17:32 2020

@author: tiago
"""

import time
import math
import locale
import numpy as np
from numpy.linalg import inv
from openpyxl import load_workbook
from itertools import permutations 

x           = [4.4, 2.2, 0, 6.6, 0, 6.6]
y           = [2.6,13.0,6.5, 10.4, 3.9, 7.8]
h           = [0.76, 1.7, 1.65, 1.17, 2.02, 1.52]
d           = []
l_rssi      = []
position    = []

path = "../Data/"

def storeData(data,sheetName):
    doc    = "methods.xlsx"
    wbk     = load_workbook(path+doc)
    page    = wbk[sheetName]
    page.append(data)
    wbk.save(path+doc)

def findDistance(a, n, rssi):
    return 10**((a - rssi)/(10*n))

def polynomial(rssi, x0, x1, x2, x3, x4):
    return x0 + x1*rssi + x2*rssi**2 + x3*rssi**3 + x4*rssi**4

# get position and RSSI from excel
def getTestData(doc):
    wbk     = load_workbook(path+doc)
    sheet   = wbk["Average"]
    cells   = sheet['A2': 'H19']
    
    for k in range(6):
        l_rssi.append([])
    
    i = -1
    locale.setlocale(locale.LC_ALL, 'en_US.UTF-8')
    for pos1, pos2, c1, c2, c3, c4, c5, c6 in cells:
        i = i + 1
        position.append([])
        position[i].append(locale.atof(pos1.value))
        position[i].append(locale.atof(pos2.value))
        l_rssi[0].append(locale.atof(c1.value))
        l_rssi[1].append(locale.atof(c2.value))
        l_rssi[2].append(locale.atof(c3.value))
        l_rssi[3].append(locale.atof(c4.value))
        l_rssi[4].append(locale.atof(c5.value))
        l_rssi[5].append(locale.atof(c6.value))
    
def polyRegression(rssi):    
    
    d.append(polynomial(rssi[0],1.90820068e+03,1.39916821e+02,3.80413206e+00,4.54250040e-02,2.01480990e-04))
    d.append(polynomial(rssi[1],8.03927288e+01,7.86690146e+00,2.73889179e-01,3.94925950e-03,2.06783022e-05))
    d.append(polynomial(rssi[2],-2.36934558e+01,-1.46667668e+00,-2.52040300e-02,-1.55811131e-04,-1.29126259e-07))
    d.append(polynomial(rssi[3],3.42350116e+01,1.80886737e+00,2.01219351e-02,-2.59560664e-04,-3.61977580e-06))
    d.append(polynomial(rssi[4],5.41055459e+02,4.39397008e+01,1.31063670e+00,1.69009265e-02,7.98114859e-05))
    d.append(polynomial(rssi[5],-2.30666363e+01,-3.59322884e+00,-1.54346312e-01,-2.70287489e-03,-1.65517534e-05))

def hyperbolicAlgorithm1(rssi):
    
    polyRegression(rssi)
    
    A = np.matrix([[ 2*(x[2] - x[5]), 2*(y[2] - y[5]) ], 
                   [ 2*(x[2] - x[1]), 2*(y[2] - y[1]) ],
                   [ 2*(x[2] - x[3]), 2*(y[2] - y[3]) ],
                   [ 2*(x[2] - x[4]), 2*(y[2] - y[4]) ]
                   #[ 2*(x[4] - x[4]), 2*(y[5] - y[4]) ]
                  ], dtype=np.float64)
    b= np.matrix([
                [d[5]**2 - d[2]**2 - x[5]**2 - y[5]**2 + x[2]**2 + y[2]**2],
                [d[1]**2 - d[2]**2 - x[1]**2 - y[1]**2 + x[2]**2 + y[2]**2],
                [d[3]**2 - d[2]**2 - x[3]**2 - y[3]**2 + x[2]**2 + y[2]**2],
                [d[4]**2 - d[2]**2 - x[4]**2 - y[4]**2 + x[2]**2 + y[2]**2],
                #[d[4]**2 - d[4]**2 - x[4]**2 - y[4]**2 + x[4]**2 + y[4]**2]
                ], dtype=np.float64)
    
    I = np.identity(2, dtype = float)
    a = 0.1
    
    AT  = A.transpose()
    AAT = AT.dot(A)
    d.clear()
    try:
        var = (inv(AAT - a*I)*AT*b).tolist()
        return [var[0][0], var[1][0]]
    except:
        print("Could not solve since matrix has no inverse.")
        return -1
        #2.910068508041401

def main():
    
    getTestData('testPoints.xlsx')
    n = len(position)

    count = 0
    log = open("Log.txt", "w+")
    minutes = 60
    start = time.time()
    while(True):
        count = count + 1
        log = open("Log.txt", "a")
        for i in range(n):
            rssi = [l_rssi[0][i],l_rssi[1][i],l_rssi[2][i],l_rssi[3][i],l_rssi[4][i],l_rssi[5][i]]
            hyperbolicAlgorithm1(rssi)
        duration = time.time() - start
        if duration - minutes > 0:
            log.write(str(minutes/60) + " " + str(count) + "\n")
            log.close()
            minutes = minutes + 60
        
if __name__== "__main__":
        main()