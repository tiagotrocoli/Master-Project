#!/usr/bin/env python2
# -*- coding: utf-8 -*-
"""
Created on Tue Feb 18 17:21:57 2020

@author: mister-c
"""

from sklearn.neural_network import MLPRegressor
from sklearn import svm
import sys
import time
import math
import numpy as np
import locale
from openpyxl import load_workbook
from scipy.optimize import minimize

x           = [4.4, 2.2, 0, 6.6, 0, 6.6]
y           = [2.6,13.0,6.5, 10.4, 3.9, 7.8]
h           = [0.76, 1.7, 1.65, 1.17, 2.02, 1.52]
distance    = []
k           = 0

path = "../Data/"

def mse(var):
    sum = 0
    for i in range(0,4):
        sum = sum + ( (x[i] - var[0])**2 + (y[i] - var[1])**2 + (h[i] - 0.73)**2 - distance[i][k]**2)**2

    return sum

def getTrainingData(doc, page, ini, end):
    wbk     = load_workbook(path+doc)
    sheet   = wbk[page]
    cells   = sheet[ini : end]
    
    Y = []
    X = []

    locale.setlocale(locale.LC_ALL, 'en_US.UTF-8')
    for distance, c1, c2, c3, c4, c5, c6, c7, c8, c9, c10 in cells:
        for k in range(10):
            Y.append(distance.value)
        X.append(c1.value) 
        X.append(c2.value)
        X.append(c3.value) 
        X.append(c4.value) 
        X.append(c5.value)
        X.append(c6.value)
        X.append(c7.value)
        X.append(c8.value)
        X.append(c9.value)
        X.append(c10.value)
    
    return X, Y

def getPosition(ini,end):
    wbk     = load_workbook(path+"testPoints.xlsx")
    sheet   = wbk["distances"]
    cells   = sheet[ini : end]
    
    position = []

    locale.setlocale(locale.LC_ALL, 'en_US.UTF-8')
    for x, y in cells:
        position.append([x.value,y.value])
    
    return position

def getTestData(doc, page, ini, end):
    wbk     = load_workbook(path+doc)
    sheet   = wbk[page]
    cells   = sheet[ini : end]
    
    Y = []
    X = []

    locale.setlocale(locale.LC_ALL, 'en_GB.UTF-8')
    for distance, c1 in cells:
        Y.append(distance.value)
        X.append(float(c1.value)) 
    
    return X, Y

def regressionNN(sheet):
    
    # get training data
    X, Y = getTrainingData('dataBase.xlsx',sheet ,"A2", "K44")
    training_X = np.array(X).reshape(-1,1)
    training_Y = np.array(Y)
    # Do regression, logistic is better
    reg = MLPRegressor(solver='lbfgs', activation='logistic', hidden_layer_sizes=(10,))
    reg.fit(training_X, training_Y)
    
    return reg
    
def estimateDistance(sheet, reg):
    
    distances = []
    
    # get test data
    X, Y = getTestData('testPoints.xlsx', sheet,  "A2", "B19")
    test_X = np.array(X).reshape(-1,1)
    test_Y = np.array(Y).reshape(-1,1)   
    
    for i in range(0,18):
        distances.extend( reg.predict(test_X[i].reshape(-1,1)).tolist() )
    
    return distances

def storeData(data,sheetName):
    doc    = "methods.xlsx"
    wbk     = load_workbook(path+doc)
    page    = wbk[sheetName]
    page.append(data)
    wbk.save(path+doc)

def main():
    
    global k
    
    # Apply Neural Network for all sensor data
    reg1 = regressionNN("TiagoLocalizacao1")
    reg2 = regressionNN("TiagoLocalizacao2")
    #reg0 = regressionNN("TiagoLocalizacao0")
    reg3 = regressionNN("TiagoLocalizacao3")
    reg4 = regressionNN("TiagoLocalizacao4")
    
    # Estimate distances for all sensor based on NN
    start = time.time()
    distance.append( estimateDistance("TiagoLocalizacao1", reg1) )
    distance.append( estimateDistance("TiagoLocalizacao2", reg2) )
    #distance.append( estimateDistance("TiagoLocalizacao0", reg0) )
    distance.append( estimateDistance("TiagoLocalizacao3", reg3) )
    distance.append( estimateDistance("TiagoLocalizacao4", reg4) )
    setupTime = (time.time() - start)/18.0
    
    position  = getPosition("A2", "B19")
    
    for k in range(18):
        x0  = np.array([1.0, 1.0])
        
        # estimate the position
        start = time.time()
        res = minimize(mse, x0, method='BFGS', options={'gtol': 1e-8})
        duration = time.time() - start
        
        # calculate accuracy
        accuracy = math.sqrt( (position[k][0] - res.x[0])**2 + (position[k][1] - res.x[1])**2 )
        
        # store parameters 
        data = [position[k][0], position[k][1]]
        data.extend([round(res.x[0],2), round(res.x[1],2)])
        data.append(setupTime+duration)
        data.append(accuracy)
        
        # send them to xlxs document
        storeData(data,"MultNN")
        
    
if __name__== "__main__":
        main()