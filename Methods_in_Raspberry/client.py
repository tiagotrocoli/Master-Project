#!/usr/bin/env python2
# -*- coding: utf-8 -*-
"""
Created on Thu Feb 20 13:38:37 2020

@author: mister-c
"""

import sys
import socket
import time
import locale
import pickle
from openpyxl import load_workbook

path = "../Data/"

def getData(doc, page, ini, end):
    wbk     = load_workbook(path+doc)
    sheet   = wbk[page]
    cells   = sheet[ini : end]
    
    position = []
    row = []

    i = -1
    locale.setlocale(locale.LC_ALL, 'en_GB.UTF-8')
    for pos1, pos2, c1, c2, c3, c4, c5, c6 in cells:
        i = i + 1
        position.append([])
        position[i].append(locale.atof(pos1.value))
        position[i].append(locale.atof(pos2.value))
        row.append([locale.atof(c1.value), locale.atof(c2.value), locale.atof(c3.value), locale.atof(c4.value), locale.atof(c5.value)])
    
    return row, position

def storeData(data,sheetName):
    doc    = "energy.xlsx"
    wbk     = load_workbook(doc)
    page    = wbk[sheetName]
    page.append(data)
    wbk.save(doc)

def main():
    
    ipv4 = sys.argv[1]
    port = int(sys.argv[2])
    
    mySocket = socket.socket( socket.AF_INET, socket.SOCK_STREAM ) 
    mySocket.connect( (ipv4, port) )
    
    rssi_test, test_points = getData('testPoints.xlsx', "Average",  "A2", "H19")
    n = len(test_points)
    
    count = 0
    minutes = 60
    start = time.time()
    
    while(True):
        
        count = count + 1
        
        for i in range(0,n):
                            
            data = pickle.dumps(rssi_test[i])
            mySocket.send(data)
            recv_data = mySocket.recv( 1024 )
            recv_data = pickle.loads(recv_data)
        
            recv_data[4] = 1
            #print(recv_data)
        duration = time.time() - start
        if duration - minutes > 0:
            data = []
            data.append(minutes/60)
            data.append(count)
            storeData(data, "counter")
            minutes = minutes + 60
        
    mySocket.close()

if __name__== "__main__":
        main()