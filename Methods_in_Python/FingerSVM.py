#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Fri Mar 13 22:41:34 2020

@author: tiago
"""

from sklearn.neural_network import MLPRegressor
from sklearn import svm
import sys
import time
import math
import numpy as np
import locale
from openpyxl import load_workbook
from sklearn import svm

x           = [4.4, 2.2, 0, 6.6, 0, 6.6]
y           = [2.6,13.0,6.5, 10.4, 3.9, 7.8]
h           = [0.76, 1.7, 1.65, 1.17, 2.02, 1.52]

path = "../Data/"

def storeData(data,sheetName):
    path    = "methods.xlsx"
    wbk     = load_workbook(path)
    page    = wbk[sheetName]
    page.append(data)
    wbk.save(path)

def getData(doc, page, ini, end):
    wbk     = load_workbook(path+doc)
    sheet   = wbk[page]
    cells   = sheet[ini : end]
    
    position = []
    row = []

    i = -1
    locale.setlocale(locale.LC_ALL, 'en_US.UTF-8')
    for pos1, pos2, c1, c2, c3, c4, c5 in cells:
        i = i + 1
        position.append([])
        position[i].append(locale.atof(pos1.value))
        position[i].append(locale.atof(pos2.value))
        row.append([locale.atof(c1.value), locale.atof(c2.value), locale.atof(c3.value), locale.atof(c4.value), locale.atof(c5.value)])
    
    return row, position

def main():
    
    base_rssi, base_pos = getData('dataBase.xlsx',"Average" ,"A2", "G44")
    test_rssi, test_pos = getData('testPoints.xlsx', "Average",  "A2", "G19")
    training_X = np.array(base_rssi)
    training_Y = np.array(base_pos)
    test_X = np.array(test_rssi)
    clf = svm.SVR(kernel='poly')
    print("oi")
    clf.fit(training_X, training_Y[:,1])
    print("oi")
    #reg = MLPRegressor(solver='lbfgs', activation='identity', hidden_layer_sizes=(100,))
    #reg.fit(training_X, training_Y)
    for i in range(0,18):
        #pred_y_test = reg.predict(test_X[i].reshape(1,-1))
        pred_y_test = clf.predict(test_X[i].reshape(1,-1))
        print(pred_y_test)
        
if __name__== "__main__":
        main()