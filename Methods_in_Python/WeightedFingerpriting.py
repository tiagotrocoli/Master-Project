#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Sun Feb  9 18:54:06 2020

@author: tiago
"""


import sys
import time
import math
import locale
import numpy as np
from openpyxl import load_workbook

#x           = [4.4, 2.2, 0, 6.6, 0, 6.6]
#y           = [2.6,13.0,6.5, 10.4, 3.9, 7.8]
#h           = [0.76, 1.7, 1.65, 1.17, 2.02, 1.52]


path = "../Data/"

def storeData(data,sheetName):
    doc    = "methods.xlsx"
    wbk     = load_workbook(path+doc)
    page    = wbk[sheetName]
    page.append(data)
    wbk.save(path+doc)

def getData(doc, page, ini, end):
    wbk     = load_workbook(path+doc)
    sheet   = wbk[page]
    cells   = sheet[ini : end]
    
    position = []
    row = []

    i = -1
    locale.setlocale(locale.LC_ALL, 'en_US.UTF-8')
    for pos1, pos2, c1, c2, c3, c4, c5, c6 in cells:
        i = i + 1
        position.append([])
        position[i].append(locale.atof(pos1.value))
        position[i].append(locale.atof(pos2.value))
        row.append([locale.atof(c1.value), locale.atof(c2.value), locale.atof(c3.value), locale.atof(c4.value), locale.atof(c5.value)])
    
    return row, position

def getData2(doc, page, ini, end):
    wbk     = load_workbook(path+doc)
    sheet   = wbk[page]
    cells   = sheet[ini : end]
    
    position = []
    row = []

    i = -1
    locale.setlocale(locale.LC_ALL, 'en_US.UTF-8')
    for pos1, pos2, c1, c2, c3, c4, c5, c6 in cells:
        i = i + 1
        position.append([])
        position[i].append(pos1.value)
        position[i].append(pos2.value)
        row.append([c1.value, c2.value, c3.value, c4.value, c5.value, c6.value])
    return row, position
    
def main():
    
    #n = int(sys.argv[1])
    comb = [4,3,2,1]
    n = len(comb)
    # get list of rssi and list of position, respectivelly
    l_base, base_pos = getData2('dataBase.xlsx',"Experiment2" ,"A2", "H38")
    l_test, test_pos = getData2('testPoints.xlsx', "Experiment2",  "A2", "H18")
    
    k = -1
    avg = 0
    w0 = 10**(-8)
    std_error = []
    for point in l_test:
        k = k + 1
        # execute Fingerpriting and calculate its processing time
        cost = []
        start = time.time()
        for i in range(37):
            total = 0
            for r in range(n):
                total = total + (point[comb[r]] - l_base[i][comb[r]])**2
            cost.append(total)
        # return a list of index of sorted cost
        index = np.argsort(cost)
        # set the weight
        weight = 0
        for j in range(0,n):
            weight = weight + 1.0/(cost[index[j]]+w0)
        # calculate x,y via weighted cost
        x = y = 0
        for j in range(0,n):
            x = x + ((1.0/(cost[index[j]]+w0))/weight)*base_pos[index[j]][0]
            y = y + ((1.0/(cost[index[j]]+w0))/weight)*base_pos[index[j]][1]
        duration = time.time() - start
        estimate = [round(x, 2),round(y, 2)]
        # calculate error
        error = math.sqrt((estimate[0] - test_pos[k][0])**2 + (estimate[1] - test_pos[k][1])**2)
        std_error.append(error)
        # put them together
        data = [test_pos[k][0], test_pos[k][1]]
        data.extend(estimate)
        data.append(duration)
        data.append(round(error,2))
        avg = avg + error
        # store in xlsx
        storeData(data, "WeightedFingerKnn_exp2")
    print(avg/17.0, np.std(std_error))
        
        
if __name__== "__main__":
        main()