#!/usr/bin/env python2
# -*- coding: utf-8 -*-
"""
Created on Mon Feb 10 17:04:36 2020

@author: mister-c
"""

from sklearn.neural_network import MLPRegressor
from sklearn import svm
import sys
import time
import math
import numpy as np
import locale
from openpyxl import load_workbook
from sklearn import svm

x           = [4.4, 2.2, 0, 6.6, 0, 6.6]
y           = [2.6,13.0,6.5, 10.4, 3.9, 7.8]
h           = [0.76, 1.7, 1.65, 1.17, 2.02, 1.52]

path = "../Data/"

def storeData(data,sheetName):
    path    = "methods.xlsx"
    wbk     = load_workbook(path)
    page    = wbk[sheetName]
    page.append(data)
    wbk.save(path)

def getData(doc, page, ini, end):
    wbk     = load_workbook(path+doc)
    sheet   = wbk[page]
    cells   = sheet[ini : end]
    
    position = []
    row = []

    i = -1
    locale.setlocale(locale.LC_ALL, 'en_US.UTF-8')
    for pos1, pos2, c1, c2, c3, c4, c5 in cells:
        i = i + 1
        position.append([])
        position[i].append(locale.atof(pos1.value))
        position[i].append(locale.atof(pos2.value))
        row.append([locale.atof(c1.value), locale.atof(c2.value), locale.atof(c3.value), locale.atof(c4.value), locale.atof(c5.value)])
    
    return row, position

def getData2(doc, page, ini, end):
    wbk     = load_workbook(path+doc)
    sheet   = wbk[page]
    cells   = sheet[ini : end]
    
    position = []
    row = []

    i = -1
    locale.setlocale(locale.LC_ALL, 'en_US.UTF-8')
    for pos1, pos2, c1, c2, c3, c4, c5, c6 in cells:
        i = i + 1
        position.append([])
        position[i].append(pos1.value)
        position[i].append(pos2.value)
        row.append([c1.value, c2.value, c3.value, c4.value, c5.value, c6.value])

    return row, position

def main():
    
    avg = 0
    base_rssi, base_pos = getData2('dataBase.xlsx',"Experiment2" ,"A2", "H38")
    test_rssi, test_pos = getData2('testPoints.xlsx', "Average",  "A2", "H18")
    test_pos = np.array(test_pos, dtype = np.float32)
    training_X = np.array(base_rssi, dtype = np.float32)
    training_Y = np.array(base_pos, dtype = np.float32)
    test_X = np.array(test_rssi, dtype = np.float32)
    reg = MLPRegressor(solver='lbfgs', activation='identity', hidden_layer_sizes=(10,))
    reg.fit(training_X, training_Y)
    for i in range(0,17):
        start = time.time()
        pred_y_test = reg.predict(test_X[i].reshape(1,-1))
        duration = time.time() - start
        accuracy = np.sqrt((pred_y_test[0][0] - test_pos[i][0])**2 + (pred_y_test[0][1] - test_pos[i][1])**2)
        data = []
        data.extend(test_pos[i])
        data.extend(pred_y_test[0])
        data.append(duration)
        data.append(accuracy)
        avg = avg + accuracy
        print(data)
        #storeData(data,"FingerNN")
    print(avg/17.0)
if __name__== "__main__":
        main()
