#!/usr/bin/env python2
# -*- coding: utf-8 -*-
"""
Created on Tue Feb  4 13:27:22 2020

@author: mister-c
"""

import time
import math
import locale
import numpy as np
from openpyxl import load_workbook
from scipy.optimize import minimize

#x           = [4.4, 2.2, 0, 6.6, 0, 6.6]
#y           = [2.6,13.0,6.5, 10.4, 3.9, 7.8]
#h           = [0.76, 1.7, 1.65, 1.17, 2.02, 1.52]
d           = [] # list of distances
l_rssi      = []
position    = []


x = [-4.0, 4.0, -4.6, -2.0, -2.74, 2.0]
y = [4.6, 2.4, 1.84, 5.52, 0.92, 0.8]

path = "../Data/"

def mse(var):
    sum = 0
    
    sum = sum + ( (x[0] - var[0])**2 + (y[0] - var[1])**2 + (h[0] - 0.73)**2 - d[0]**2)**2
    sum = sum + ( (x[1] - var[0])**2 + (y[1] - var[1])**2 + (h[1] - 0.73)**2 - d[1]**2)**2
    sum = sum + ( (x[2] - var[0])**2 + (y[2] - var[1])**2 + (h[2] - 0.73)**2 - d[2]**2)**2
    sum = sum + ( (x[3] - var[0])**2 + (y[3] - var[1])**2 + (h[3] - 0.73)**2 - d[3]**2)**2
    sum = sum + ( (x[4] - var[0])**2 + (y[4] - var[1])**2 + (h[4] - 0.73)**2 - d[4]**2)**2
    sum = sum + ( (x[5] - var[0])**2 + (y[5] - var[1])**2 + (h[5] - 0.73)**2 - d[5]**2)**2
    
    return sum

def mse2(var):
    sum = 0
    
    sum = sum + ( (x[0] - var[0])**2 + (y[0] - var[1])**2 - d[0]**2)**2
    sum = sum + ( (x[1] - var[0])**2 + (y[1] - var[1])**2 - d[1]**2)**2
    sum = sum + ( (x[2] - var[0])**2 + (y[2] - var[1])**2 - d[2]**2)**2
    #sum = sum + ( (x[3] - var[0])**2 + (y[3] - var[1])**2 - d[3]**2)**2
    #sum = sum + ( (x[4] - var[0])**2 + (y[4] - var[1])**2 - d[4]**2)**2
    #sum = sum + ( (x[5] - var[0])**2 + (y[5] - var[1])**2 - d[5]**2)**2
    
    return sum

def storeData(data,sheetName):
    doc    = "methods.xlsx"
    wbk     = load_workbook(path+doc)
    page    = wbk[sheetName]
    page.append(data)
    wbk.save(path+doc)

def findDistance(a, n, rssi):
    return 10**((a - rssi)/(10*n))

def polynomial(rssi, x0, x1, x2, x3, x4):
    return x0 + x1*rssi + x2*rssi**2 + x3*rssi**3 + x4*rssi**4

# get position and RSSI from excel
def getTestData(doc, page, ini, end):
    wbk     = load_workbook(path+doc)
    sheet   = wbk[page]
    cells   = sheet[ini: end]
    
    for k in range(6):
        l_rssi.append([])
    
    i = -1
    locale.setlocale(locale.LC_ALL, 'en_GB.UTF-8')
    for pos1, pos2, c1, c2, c3, c4, c5, c6 in cells:
        i = i + 1
        position.append([])
        position[i].append(pos1.value)
        position[i].append(pos2.value)
        l_rssi[0].append(c1.value)
        l_rssi[1].append(c2.value)
        l_rssi[2].append(c3.value)
        l_rssi[3].append(c4.value)
        l_rssi[4].append(c5.value)
        l_rssi[5].append(c6.value)

def lognomal(rssi):

    d.append(findDistance(-23.1615, 4.3478, rssi[0]) )
    d.append(findDistance(-2.4133, 5.7772, rssi[1]))
    d.append(findDistance(52.6093, 14.6044, rssi[2]))
    d.append(findDistance(0.2455, 7.4626, rssi[3]))
    d.append(findDistance(38.7733, 11.7859,rssi[4]))
    d.append(findDistance(49.1173, 13.8565, rssi[5]))
    
def lognomal_exp2(rssi):

    d.append(findDistance(8.4955, 9.4119, rssi[0]) )
    d.append(findDistance(-1.8347, 7.3338, rssi[1]))
    d.append(findDistance(-1.4735, 7.5332, rssi[2]))
    d.append(findDistance(-21.3316, 5.9053, rssi[3]))
    d.append(findDistance(0.2642, 9.749,rssi[4]))
    d.append(findDistance(-18.7628, 6.4427, rssi[5]))
    
def polyRegression(rssi):    
    
    d.append(polynomial(rssi[0],1.90820068e+03,1.39916821e+02,3.80413206e+00,4.54250040e-02,2.01480990e-04))
    d.append(polynomial(rssi[1],8.03927288e+01,7.86690146e+00,2.73889179e-01,3.94925950e-03,2.06783022e-05))
    d.append(polynomial(rssi[2],-2.36934558e+01,-1.46667668e+00,-2.52040300e-02,-1.55811131e-04,-1.29126259e-07))
    d.append(polynomial(rssi[3],3.42350116e+01,1.80886737e+00,2.01219351e-02,-2.59560664e-04,-3.61977580e-06))
    d.append(polynomial(rssi[4],5.41055459e+02,4.39397008e+01,1.31063670e+00,1.69009265e-02,7.98114859e-05))
    d.append(polynomial(rssi[5],-2.30666363e+01,-3.59322884e+00,-1.54346312e-01,-2.70287489e-03,-1.65517534e-05))
    
def circularAlgorithm(rssi):
    
    lognomal_exp2(rssi)
    
    #print(rssi)
    #print(d)
    x0  = np.array([1.0, 1.0])
    res = minimize(mse2, x0, method='BFGS', options={'gtol': 1e-8})
    d.clear()
    
    return [res.x[0], res.x[1]]
    

def main():
    
    getTestData('testPoints.xlsx', "Average_exp2", "A2", "H18")
    n = len(position)
    avg = 0
    
    for i in range(n):
        rssi = [l_rssi[0][i],l_rssi[1][i],l_rssi[2][i],l_rssi[3][i],l_rssi[4][i],l_rssi[5][i]]
        #calculate duration
        start = time.time()
        res = circularAlgorithm(rssi)
        duration = time.time() - start
        # calculate precision
        accuracy = math.sqrt( (position[i][0] - res[0])**2 + (position[i][1] - res[1])**2 )
        # put them together
        data = [position[i][0], position[i][1]]
        data.extend([round(res[0],8), round(res[1],8)])
        data.append(round(duration,8))
        data.append(round(accuracy,8))
        print (','.join(str(x) for x in data))
        avg = avg + accuracy
        # store in xlsx
        #storeData(data,"CirculaAlgo_exp2")
    print(avg/17.0)
    
if __name__== "__main__":
        main()
