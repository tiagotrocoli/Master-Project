#!/usr/bin/env python2
# -*- coding: utf-8 -*-
"""
Created on Tue Feb  4 13:45:22 2020

@author: mister-c
"""

import time
import math
import locale
import numpy as np
from numpy.linalg import inv
from openpyxl import load_workbook
from itertools import permutations 

#x           = [4.4, 2.2, 0, 6.6, 0, 6.6]
#y           = [2.6,13.0,6.5, 10.4, 3.9, 7.8]
#h           = [0.76, 1.7, 1.65, 1.17, 2.02, 1.52]
x = [-4.0, 4.0, -4.6, -2.0, -2.74, 2.0]
y = [4.6, 2.4, 1.84, 5.52, 0.92, 0.8]

d           = []
l_rssi      = []
position    = []

path = "../Data/"

def storeData(data,sheetName):
    doc    = "methods.xlsx"
    wbk     = load_workbook(path+doc)
    page    = wbk[sheetName]
    page.append(data)
    wbk.save(path+doc)

def findDistance(a, n, rssi):
    return 10**((a - rssi)/(10*n))

def polynomial(rssi, x0, x1, x2, x3, x4):
    return x0 + x1*rssi + x2*rssi**2 + x3*rssi**3 + x4*rssi**4

def getTestData2(doc, page, ini, end):
    wbk     = load_workbook(path+doc)
    sheet   = wbk[page]
    cells   = sheet[ini: end]
    
    for k in range(6):
        l_rssi.append([])
    
    i = -1
    locale.setlocale(locale.LC_ALL, 'en_GB.UTF-8')
    for pos1, pos2, c1, c2, c3, c4, c5, c6 in cells:
        i = i + 1
        position.append([])
        position[i].append(pos1.value)
        position[i].append(pos2.value)
        l_rssi[0].append(c1.value)
        l_rssi[1].append(c2.value)
        l_rssi[2].append(c3.value)
        l_rssi[3].append(c4.value)
        l_rssi[4].append(c5.value)
        l_rssi[5].append(c6.value)

# get position and RSSI from excel
def getTestData(doc):
    wbk     = load_workbook(path+doc)
    sheet   = wbk["Average"]
    cells   = sheet['A2': 'H19']
    
    for k in range(6):
        l_rssi.append([])
    
    i = -1
    locale.setlocale(locale.LC_ALL, 'en_GB.UTF-8')
    for pos1, pos2, c1, c2, c3, c4, c5, c6 in cells:
        i = i + 1
        position.append([])
        position[i].append(locale.atof(pos1.value))
        position[i].append(locale.atof(pos2.value))
        l_rssi[0].append(locale.atof(c1.value))
        l_rssi[1].append(locale.atof(c2.value))
        l_rssi[2].append(locale.atof(c3.value))
        l_rssi[3].append(locale.atof(c4.value))
        l_rssi[4].append(locale.atof(c5.value))
        l_rssi[5].append(locale.atof(c6.value))

def lognomal_exp2(rssi):

    d.append(findDistance(8.4955, 9.4119, rssi[0]) )
    d.append(findDistance(-1.8347, 7.3338, rssi[1]))
    d.append(findDistance(-1.4735, 7.5332, rssi[2]))
    d.append(findDistance(-21.3316, 5.9053, rssi[3]))
    d.append(findDistance(0.2642, 9.749,rssi[4]))
    d.append(findDistance(-18.7628, 6.4427, rssi[5]))

def lognomal(rssi):

    d.append(findDistance(-23.1615, 4.3478, rssi[0]) )
    d.append(findDistance(-2.4133, 5.7772, rssi[1]))
    d.append(findDistance(52.6093, 14.6044, rssi[2]))
    d.append(findDistance(0.2455, 7.4626, rssi[3]))
    d.append(findDistance(38.7733, 11.7859,rssi[4]))
    d.append(findDistance(49.1173, 13.8565, rssi[5]))
    
def polyRegression(rssi):    
    
    d.append(polynomial(rssi[0],1.90820068e+03,1.39916821e+02,3.80413206e+00,4.54250040e-02,2.01480990e-04))
    d.append(polynomial(rssi[1],8.03927288e+01,7.86690146e+00,2.73889179e-01,3.94925950e-03,2.06783022e-05))
    d.append(polynomial(rssi[2],-2.36934558e+01,-1.46667668e+00,-2.52040300e-02,-1.55811131e-04,-1.29126259e-07))
    d.append(polynomial(rssi[3],3.42350116e+01,1.80886737e+00,2.01219351e-02,-2.59560664e-04,-3.61977580e-06))
    d.append(polynomial(rssi[4],5.41055459e+02,4.39397008e+01,1.31063670e+00,1.69009265e-02,7.98114859e-05))
    d.append(polynomial(rssi[5],-2.30666363e+01,-3.59322884e+00,-1.54346312e-01,-2.70287489e-03,-1.65517534e-05))
   
    
def polynomial_exp2(rssi):    
    
    d.append(polynomial(rssi[0],-1.04677560e+02,-8.43418050e+00,-2.41329137e-01,-2.99784603e-03,-1.35071060e-05))
    d.append(polynomial(rssi[1],6.75010158e+01,5.42156108e+00,1.56931129e-01,1.85139117e-03,7.65654151e-06))
    d.append(polynomial(rssi[2],2.37812739e+02,1.91631461e+01,5.61308656e-01,7.02168500e-03,3.19705302e-05))
    d.append(polynomial(rssi[3],1.66812553e+02,1.30910570e+01,3.70048492e-01,4.44918165e-03,1.94958486e-05))
    d.append(polynomial(rssi[4],1.16571208e+02,1.01184914e+01,3.12116567e-01,4.04167347e-03,1.88556599e-05))
    d.append(polynomial(rssi[5],1.99163123e+01,1.78888185e+00,5.49568407e-02,6.48346195e-04,2.66866033e-06))

def hyperbolicAlgorithm(rssi):
    
    polynomial_exp2(rssi)
    
    A = np.matrix([[ 2*(x[3] - x[4]), 2*(y[3] - y[4]) ], 
                   [ 2*(x[3] - x[5]), 2*(y[3] - y[5]) ],
                   #[ 2*(x[3] - x[4]), 2*(y[3] - y[3]) ],
                   #[ 2*(x[3] - x[4]), 2*(y[3] - y[4]) ]
                   #[ 2*(x[4] - x[4]), 2*(y[5] - y[4]) ]
                  ], dtype=np.float64)
    b= np.matrix([
                [d[4]**2 - d[3]**2 - x[5]**2 - y[4]**2 + x[3]**2 + y[3]**2],
                [d[5]**2 - d[3]**2 - x[5]**2 - y[5]**2 + x[3]**2 + y[3]**2],
                #[d[3]**2 - d[3]**2 - x[3]**2 - y[3]**2 + x[3]**2 + y[3]**2],
                #[d[4]**2 - d[3]**2 - x[4]**2 - y[4]**2 + x[3]**2 + y[3]**2],
                #[d[4]**2 - d[4]**2 - x[4]**2 - y[4]**2 + x[4]**2 + y[4]**2]
                ], dtype=np.float64)
    
    AT  = A.transpose()
    AAT = AT.dot(A)
    d.clear()
    try:
        var = (inv(AAT)*AT*b).tolist()
        return [var[0][0], var[1][0]]
    except:
        print("Could not solve since matrix has no inverse.")
        return -1
        #3.4760209715908

def hyperbolicAlgorithm1(rssi, comb):
    
    polynomial_exp2(rssi)
    
    anch = comb[0]
    index1 = comb[1]
    index2 = comb[2]
    #index3 = comb[3]
    #index4 = comb[4]
    
    A = np.matrix([[ 2*(x[anch] - x[index1]), 2*(y[anch] - y[index1]) ], 
                   [ 2*(x[anch] - x[index2]), 2*(y[anch] - y[index2]) ],
                   #[ 2*(x[anch] - x[index3]), 2*(y[anch] - y[index3]) ],
                   #[ 2*(x[anch] - x[index4]), 2*(y[anch] - y[index4]) ]
                   #[ 2*(x[4] - x[4]), 2*(y[5] - y[4]) ]
                  ], dtype=np.float64)
    b= np.matrix([
                [d[index1]**2 - d[anch]**2 - x[index1]**2 - y[index1]**2 + x[anch]**2 + y[anch]**2],
                [d[index2]**2 - d[anch]**2 - x[index2]**2 - y[index2]**2 + x[anch]**2 + y[anch]**2],
                #[d[index3]**2 - d[anch]**2 - x[index3]**2 - y[index3]**2 + x[anch]**2 + y[anch]**2],
                #[d[index4]**2 - d[anch]**2 - x[index4]**2 - y[index4]**2 + x[anch]**2 + y[anch]**2],
                #[d[4]**2 - d[4]**2 - x[4]**2 - y[4]**2 + x[4]**2 + y[4]**2]
                ], dtype=np.float64)
    
    I = np.identity(2, dtype = float)
    a = 0.1
    
    AT  = A.transpose()
    AAT = AT.dot(A)
    d.clear()
    try:
        var = (inv(AAT+a*I)*AT*b).tolist()
        return [var[0][0], var[1][0]]
    except:
        print("Could not solve since matrix has no inverse.")
        return -1
        #2.910068508041401

def main():
    
    getTestData2('testPoints.xlsx', "Average_exp2", "A2", "H18")
    n = len(position)

    for comb_n in range(3,6):
        combs = permutations([0,1,2,3,4,5], comb_n)
        best = 100
        best_comb = []
        for comb in combs:
            avg = 0
            std_error = []
            for i in range(n):
                rssi = [l_rssi[0][i],l_rssi[1][i],l_rssi[2][i],l_rssi[3][i],l_rssi[4][i],l_rssi[5][i]]
                #calculate duration
                start = time.time()
                res = hyperbolicAlgorithm1(rssi, comb)
                duration = time.time() - start
                # calculate precision
                error = math.sqrt( (position[i][0] - res[0])**2 + (position[i][1] - res[1])**2 )
                # put them together
                data = [round(position[i][0],8), round(position[i][1],8)]
                data.extend([round(res[0],8), round(res[1], 8)])
                data.append(round(duration,8))
                data.append(round(error,8))
                std_error.append(error)
                #print (','.join(str(x) for x in data))
                avg = avg + error
                # store in xlsx
                #storeData(data,"HyperbolicAlgoPoly_exp2")
            #print(avg/17.0, np.std(std_error), comb)
            if best > avg/17.0:
                best = avg/17.0
                best_comb = comb
        print(best, best_comb)
    
if __name__== "__main__":
        main()
