#!/usr/bin/env python2
# -*- coding: utf-8 -*-
"""
Created on Mon Feb 10 14:16:43 2020

@author: mister-c
"""

import sys
import time
import math
import locale
from openpyxl import load_workbook
from itertools import permutations 
import numpy as np

#x           = [4.4, 2.2, 0, 6.6, 0, 6.6]
#y           = [2.6,13.0,6.5, 10.4, 3.9, 7.8]
#h           = [0.76, 1.7, 1.65, 1.17, 2.02, 1.52]

x = [-4.0, 4.0, -4.6, -2.0, -2.74, 2.0]
y = [4.6, 2.4, 1.84, 5.52, 0.92, 0.8]

path = "../Data/"

def storeData(data,sheetName):
    doc    = "methods.xlsx"
    wbk     = load_workbook(path+doc)
    page    = wbk[sheetName]
    page.append(data)
    wbk.save(path+doc)

def getData2(doc, page, ini, end):
    wbk     = load_workbook(path+doc)
    sheet   = wbk[page]
    cells   = sheet[ini : end]
    
    position = []
    row = []

    i = -1
    locale.setlocale(locale.LC_ALL, 'en_US.UTF-8')
    for pos1, pos2, c1, c2, c3, c4, c5, c6 in cells:
        i = i + 1
        position.append([])
        position[i].append(pos1.value)
        position[i].append(pos2.value)
        row.append([c1.value, c2.value, c3.value, c4.value, c5.value, c6.value])

    return row, position

def getData(doc, page, ini, end):
    wbk     = load_workbook(path+doc)
    sheet   = wbk[page]
    cells   = sheet[ini : end]
    
    position = []
    row = []

    i = -1
    locale.setlocale(locale.LC_ALL, 'en_US.UTF-8')
    for pos1, pos2, c1, c2, c3, c4, c5, c6 in cells:
        i = i + 1
        position.append([])
        position[i].append(locale.atof(pos1.value))
        position[i].append(locale.atof(pos2.value))
        row.append([locale.atof(c1.value), locale.atof(c2.value), locale.atof(c3.value), locale.atof(c4.value), locale.atof(c5.value)])
        
    return row, position

def main():
    
    #n = int(sys.argv[1])
    #comb = [0,1,2]
    #n = len(comb)
    
    #combs = permutations([0,1,2,3,4,5], 4)
    combs = [[4,3,1,2]]
    n = 4
    l_base, base_pos = getData2('dataBase.xlsx',"Experiment2" ,"A2", "H38")
    l_test, test_pos = getData2('testPoints.xlsx', "Experiment2",  "A2", "H18")
    
    for comb in combs:
        k = -1
        avg = 0
        std_error = []
        for point in l_test:
            k = k + 1
            cost = []
            # execute Fingerpriting and calculate its processing time
            start = time.time()
            for i in range(37):
                total = 0
                for r in range(n):
                    total = total + (point[comb[r]] - l_base[i][comb[r]])**2
                cost.append(total)
            # sort cost and index of base_pos accordingly
            pos = [x for _,x in sorted(zip(cost,base_pos))]
            x = y = 0
            for j in range(n):
                x = x + pos[j][0]
                y = y + pos[j][1]
            x = x/(1.0*n)
            y = y/(1.0*n)
            duration = time.time() - start
            estimate = [round(x, 2),round(y, 2)]
            # calculate error
            error = math.sqrt((estimate[0] - test_pos[k][0])**2 + (estimate[1] - test_pos[k][1])**2)
            std_error.append(error)
            # put them together
            data = [test_pos[k][0], test_pos[k][1]]
            data.extend(estimate)
            data.append(duration)
            data.append(round(error,2))
            avg = avg + error
            #print(data)
            # store in xlsx
            storeData(data, "FingerKNN_exp2")
        print(avg/17.0, np.std(std_error), comb)
        
if __name__== "__main__":
        main()
